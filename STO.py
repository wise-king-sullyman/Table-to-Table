"""organize the structure type list from a DOT table into our SEI table"""

import openpyxl

def total():
    wb = openpyxl.load_workbook('S-TYPE.xlsx')
    ws = wb['Sheet1']
    ws2 = wb['Sheet2']
    rows = tuple(ws.rows)
    columns = tuple(ws.columns)
    columns2 = tuple(ws2.columns)

    rowCount = 2
    colCount = 1  
    colCount2 = 1

    for row in rows:
        for column in columns:
            currentCell = ws.cell(row=rowCount, column=colCount)           
            if currentCell.value:
                currentCellLabel = (ws.cell(row=1, column=colCount)).value
                for column in columns2: #deposit detected value in correct new column
                    targetCell = ws2.cell(row=rowCount, column=colCount2)
                    targetCellLabel = (ws2.cell(row=1, column=colCount2)).value
                    if currentCellLabel == targetCellLabel:
                        #print('foo')
                        targetCell.value = currentCell.value
                        colCount2 = 1
                        break
                    else:
                        colCount2 += 1
            else:
                colCount += 1
        colCount = 1
        rowCount += 1

    wb.save("S-TYPE_T.xlsx")

def populateRows(): #populare zeros in all empty cells of rows with at least one value in them
    wb = openpyxl.load_workbook('S-TYPE_T.xlsx')
    ws = wb['Sheet2']
    rows = tuple(ws.rows)
    columns = tuple(ws.columns)

    rowCount = 1
    colCount = 1

    for row in rows: #iterate through table
        for column in columns:
            currentCell = ws.cell(row=rowCount, column=colCount)
            if currentCell.value: #if a value is found
                colCount = 1             
                #print('foo')  
                for column in columns: #start at beggining and iterate through row
                    currentCell = ws.cell(row=rowCount, column=colCount)
                    if currentCell.value:
                        #print('baz')                
                        colCount += 1
                    else:
                        currentCell.value = 0 #replace empty cells in these rows with zero
                        colCount += 1
            else:
                #print('bar')
                colCount += 1
        colCount = 1 #reset column counter and increment row counter
        rowCount += 1

    wb.save('S-TYPE_NB.xlsx')

total()
populateRows()