import openpyxl

def organizePipes(): #create new spreadsheet with pipe length/sizes in correct order
    wb = openpyxl.load_workbook('lengths.xlsx')
    ws1 = wb['Sheet1']
    ws2 = wb['Sheet2']
    rows = tuple(ws1.rows)
    columns = tuple(ws1.columns)

    rowCount = 1
    colCount = 1
    colCount2 = 1

    for i in range(1, 70): #iterate through all pipe sizes smallest to largest
        for column in columns:
            currentCell = ws1.cell(row=rowCount, column=colCount)
            if currentCell.value == i: #if selected pipe size is found
                for row in rows:    #copy all rows beneth the pipe size statement into output sheet
                    targetCell = ws2.cell(row=rowCount, column=colCount2) 
                    targetCell.value = currentCell.value
                    rowCount += 1
                    currentCell = ws1.cell(row=rowCount, column=colCount)
                rowCount = 1
                colCount += 1
                colCount2 += 1

            else:
                colCount +=1 #iterate onto next column in base sheet if target pipe size isn't found 

        colCount = 1


    wb.save('lengths2.xlsx')

def populateRows(): #populare zeros in all empty cells of rows with at least one value in them
    wb = openpyxl.load_workbook('lengths2.xlsx')
    ws1 = wb['Sheet2']
    rows = tuple(ws1.rows)
    columns = tuple(ws1.columns)

    rowCount = 1
    colCount = 1

    for row in rows: #iterate through table
        for column in columns:
            currentCell = ws1.cell(row=rowCount, column=colCount)
            if currentCell.value: #if a value is found
                colCount = 1             
                #print('foo')  
                for column in columns: #start at beggining and iterate through row
                    currentCell = ws1.cell(row=rowCount, column=colCount)
                    if currentCell.value:
                        #print('baz')                
                        colCount += 1
                    else:
                        currentCell.value = 0 #replace empty cells in these rows with zero
                        colCount += 1
            else:
                #print('bar')
                colCount += 1
        colCount = 1 #reset column counter and increment row counter
        rowCount += 1

    wb.save('lengths3.xlsx')


organizePipes()
populateRows()